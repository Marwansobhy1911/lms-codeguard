import io
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from src.lms.models import User, Task, SessionSchedule, Submission, Attendance, AttendanceStatusEnum, ProjectGrade, get_egypt_now

TEMPLATE_FILENAME = "SFE_Grades_Calculation_Model_Updated(12).xlsx"

def build_sfe_grades_workbook(db: Session, template_path: str = None) -> io.BytesIO:
    """
    Generates the complete SFE Grades Calculation Model Excel workbook
    matching the exact structure, distribution, formulas, styles, and sheets
    of SFE_Grades_Calculation_Model_Updated(12).xlsx, dynamically populated with
    the latest student data, task submissions, session attendance, and bonus points from the LMS DB.
    """
    if template_path is None:
        # Check standard locations
        root_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        candidate_1 = os.path.join(root_dir, TEMPLATE_FILENAME)
        candidate_2 = os.path.join(os.getcwd(), TEMPLATE_FILENAME)
        if os.path.exists(candidate_1):
            template_path = candidate_1
        elif os.path.exists(candidate_2):
            template_path = candidate_2

    # Query active DB data in website database order
    students = (
        db.query(User)
        .filter(User.role.like("%student%"))
        .all()
    )
    students.sort(key=lambda u: int(u.id) if str(u.id).isdigit() else 999999)
    tasks = db.query(Task).order_by(Task.id.asc()).all()
    sessions = db.query(SessionSchedule).order_by(SessionSchedule.id.asc()).all()
    all_submissions = db.query(Submission).all()
    all_attendances = db.query(Attendance).all()
    all_project_grades = db.query(ProjectGrade).all()

    # Map user id / seat number -> user object
    user_by_id = {str(u.id): u for u in students}
    user_by_seat = {str(u.seat_number): u for u in students if u.seat_number}

    # Map student_id / seat_number -> ProjectGrade
    saved_grades_map = {}
    for pg in all_project_grades:
        saved_grades_map[str(pg.student_id)] = pg
        stu_obj = user_by_id.get(str(pg.student_id))
        if stu_obj and stu_obj.seat_number:
            saved_grades_map[str(stu_obj.seat_number)] = pg

    # Build lookup mappings
    # student_id -> {task_id: score}
    sub_map = {}
    for sub in all_submissions:
        if sub.student_id not in sub_map:
            sub_map[sub.student_id] = {}
        sub_map[sub.student_id][sub.task_id] = (sub.score if sub.score is not None else 0.0)

    # student_id -> {session_id: status_ar}
    att_map = {}
    for att in all_attendances:
        if att.student_id not in att_map:
            att_map[att.student_id] = {}
        st_val = att.status.value if hasattr(att.status, "value") else str(att.status)
        att_map[att.student_id][att.session_id] = st_val

    status_ar = {
        "present": "حاضر",
        "absent": "غائب",
        "excused": "مستأذن"
    }

    # If template file exists, load it and update data & formulas; otherwise construct from scratch
    if template_path and os.path.exists(template_path):
        wb = openpyxl.load_workbook(template_path)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active) # remove default sheet

    # 1. Update/Populate 'Source_Grades'
    _populate_source_grades_sheet(wb, students, tasks, sessions, sub_map, att_map, status_ar, saved_grades_map)

    # 2. Update/Populate 'Project_Scores'
    _populate_project_scores_sheet(wb, students, user_by_id, user_by_seat, template_path, saved_grades_map)

    # 3. Update/Populate 'Project_Teams'
    _populate_project_teams_sheet(wb, students, user_by_id, user_by_seat)

    # 4. Update/Populate 'Team_Summary'
    _populate_team_summary_sheet(wb)

    # 5. Ensure 'Instructions' sheet exists and has proper styling
    _ensure_instructions_sheet(wb)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _get_styles():
    thin_border_side = Side(style="thin", color="D9D9D9")
    thick_border_side = Side(style="medium", color="1F4E79")
    grid_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    header_fill_dark = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_fill_blue = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_fill_yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    header_fill_highlight = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    input_fill_yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    formula_fill_green = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    lookup_fill_blue = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

    header_font_white = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_font_dark = Font(name="Calibri", size=11, bold=True, color="000000")
    bold_font = Font(name="Calibri", size=11, bold=True)
    regular_font = Font(name="Calibri", size=11)
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    return {
        "grid_border": grid_border,
        "thick_border_side": thick_border_side,
        "header_fill_dark": header_fill_dark,
        "header_fill_blue": header_fill_blue,
        "header_fill_yellow": header_fill_yellow,
        "header_fill_highlight": header_fill_highlight,
        "input_fill_yellow": input_fill_yellow,
        "formula_fill_green": formula_fill_green,
        "lookup_fill_blue": lookup_fill_blue,
        "header_font_white": header_font_white,
        "header_font_dark": header_font_dark,
        "bold_font": bold_font,
        "regular_font": regular_font,
        "align_center": align_center,
        "align_left": align_left,
        "align_right": align_right,
    }


def _ensure_instructions_sheet(wb):
    styles = _get_styles()
    if "Instructions" not in wb.sheetnames:
        ws = wb.create_sheet(title="Instructions", index=0)
    else:
        ws = wb["Instructions"]

    # If sheet is empty or needs formatting
    if ws.max_row <= 1 or ws.cell(1, 1).value is None:
        ws.cell(1, 1, "نظام حساب المعدل النهائي - SFE / LMS Grades Model")
        ws.cell(1, 1).font = Font(name="Calibri", size=16, bold=True, color="1F4E79")
        ws.merge_cells("A1:F1")
        
        ws.cell(3, 1, "الأوزان والمكونات").font = styles["bold_font"]
        headers = ["المكون", "الوزن", "الدرجة الكلية", "ملاحظات"]
        for c_idx, h in enumerate(headers, 1):
            cell = ws.cell(4, c_idx, h)
            cell.font = styles["header_font_white"]
            cell.fill = styles["header_fill_dark"]
            cell.alignment = styles["align_center"]

        rows_data = [
            ("التاسكات (Tasks)", "20%", "135", "مجموع Task1 + Task2 + Task3"),
            ("الحضور (Attendance)", "30%", "sessions × 20", "عدد الحاضر × 20   (كل سيشن = 20)"),
            ("البروجيكت (Project)", "50%", "135", "فردي 30 + جماعي 80 + بونص 25")
        ]
        for r_idx, r_val in enumerate(rows_data, 5):
            for c_idx, val in enumerate(r_val, 1):
                cell = ws.cell(r_idx, c_idx, val)
                cell.font = styles["regular_font"]
                cell.alignment = styles["align_center"] if c_idx in (2, 3) else styles["align_left"]

        ws.cell(9, 1, "شروط النجاح").font = styles["bold_font"]
        ws.cell(10, 1, "1. نسبة الحضور ≥ 60% (يحصل على الساعتين فقط إذا تجاوز الحضور 60%)").font = styles["regular_font"]
        ws.cell(11, 1, "2. المعدل النهائي > 60%").font = styles["regular_font"]
        ws.cell(12, 1, "حالة النجاح = ناجح فقط إذا تحقق الشرطان معاً").font = styles["bold_font"]

        ws.cell(14, 1, "كيفية الاستخدام").font = styles["bold_font"]
        instructions = [
            "1. في شيت Project_Scores: املأ الأعمدة الصفراء (Individual / Full Project / Bonus) لكل طالب وتيم.",
            "2. شيت Project_Teams يسحب الدرجات وبيانات التاسكات والحضور تلقائياً من Project_Scores و Source_Grades.",
            "3. جميع النسب والمعدل النهائي وحالة النجاح تُحسب تلقائياً بالمعادلات.",
            "4. الحساب تلقائي لجميع السيشنات والتاسكات المدخلة في النظام.",
            "5. كل تيم مفصول ومحدد بوضوح مع اسم المشرف المسئول عن مناقشة التيم (Supporter)."
        ]
        for idx, inst in enumerate(instructions, 15):
            ws.cell(idx, 1, inst).font = styles["regular_font"]

        ws.column_dimensions["A"].width = 45
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 40


def _populate_source_grades_sheet(wb, students, tasks, sessions, sub_map, att_map, status_ar, saved_grades_map=None):
    styles = _get_styles()
    if saved_grades_map is None:
        saved_grades_map = {}

    if "Source_Grades" not in wb.sheetnames:
        ws = wb.create_sheet(title="Source_Grades")
    else:
        ws = wb["Source_Grades"]

    # Header Row
    headers = [
        "Student ID", "Student Name", "Seat Number", "Academic Level", "Program",
        "Official Email", "Personal Email", "Phone", "Assigned Supporter (TA)",
        "Task 1", "Task 2", "Task 3", "Total Tasks Score", "Bonus Points",
        "Session 1", "Session 2", "Session 3", "Session 4", "Session 5", "Session 6",
        "Attended Sessions", "Attendance Rate (%)", "Total Score (Tasks + Bonus)",
        "Final Rank Score", "Attended Count", "Total Sessions", "Final Score & Best attendance", "Attended"
    ]

    for c_idx, h in enumerate(headers, 1):
        cell = ws.cell(1, c_idx, h)
        cell.font = styles["header_font_white"]
        cell.fill = styles["header_fill_blue"]
        cell.alignment = styles["align_center"]
        cell.border = styles["grid_border"]

    # Setup Score Board Headers on Source_Grades (AC to AE)
    ws.merge_cells("AC2:AE3")
    sb_header = ws.cell(2, 29, "Score Board (Top 15)")
    sb_header.font = Font(name="Calibri", size=12, bold=True, color="000000")
    sb_header.fill = styles["header_fill_highlight"]
    sb_header.alignment = styles["align_center"]

    # Compute students' live stats to build rank list
    student_scores = []
    
    # Fill student rows (Row 2 onwards)
    for r_idx, u in enumerate(students, 2):
        u_subs = sub_map.get(u.id, {})
        u_atts = att_map.get(u.id, {})
        
        # Tasks 1..3
        t1 = u_subs.get(1, 0.0) if len(tasks) >= 1 else 0.0
        t2 = u_subs.get(2, 0.0) if len(tasks) >= 2 else 0.0
        t3 = u_subs.get(3, 0.0) if len(tasks) >= 3 else 0.0
        bonus = u.bonus_points or 0.0
        
        # Sessions 1..6 (Session 6 reflects Project Discussion Attendance)
        s1 = status_ar.get(u_atts.get(1, "absent"), "غائب") if len(sessions) >= 1 else "غائب"
        s2 = status_ar.get(u_atts.get(2, "absent"), "غائب") if len(sessions) >= 2 else "غائب"
        s3 = status_ar.get(u_atts.get(3, "absent"), "غائب") if len(sessions) >= 3 else "غائب"
        s4 = status_ar.get(u_atts.get(4, "absent"), "غائب") if len(sessions) >= 4 else "غائب"
        s5 = status_ar.get(u_atts.get(5, "absent"), "غائب") if len(sessions) >= 5 else "غائب"

        # Check Project Discussion Attendance for Session 6
        pg = saved_grades_map.get(str(u.id)) or (saved_grades_map.get(str(u.seat_number)) if u.seat_number else None) if saved_grades_map else None
        if pg and pg.attendance:
            s6 = "حاضر"
        elif u_atts.get(6) in ("present", "حاضر"):
            s6 = "حاضر"
        else:
            s6 = status_ar.get(u_atts.get(6, "absent"), "غائب") if len(sessions) >= 6 else "غائب"

        supporter_name = u.assigned_supporter.name if u.assigned_supporter else ""

        # Basic Info
        student_code = str(u.seat_number or u.id)
        ws.cell(r_idx, 1, student_code)
        ws.cell(r_idx, 2, u.name or "")
        ws.cell(r_idx, 3, str(u.seat_number or ""))
        ws.cell(r_idx, 4, u.academic_level or "")
        ws.cell(r_idx, 5, u.program or "")
        ws.cell(r_idx, 6, u.official_email or "")
        ws.cell(r_idx, 7, u.email or "")
        ws.cell(r_idx, 8, u.phone or "")
        ws.cell(r_idx, 9, supporter_name)

        # Task Scores
        ws.cell(r_idx, 10, t1)
        ws.cell(r_idx, 11, t2)
        ws.cell(r_idx, 12, t3)

        # Total Tasks Score Formula: =IF(COUNTA(J{r}:L{r})=0,0,SUM(J{r}:L{r}))
        ws.cell(r_idx, 13, f"=IF(COUNTA(J{r_idx}:L{r_idx})=0,0,SUM(J{r_idx}:L{r_idx}))")
        
        # Bonus Points
        ws.cell(r_idx, 14, bonus)

        # Sessions 1..6
        ws.cell(r_idx, 15, s1)
        ws.cell(r_idx, 16, s2)
        ws.cell(r_idx, 17, s3)
        ws.cell(r_idx, 18, s4)
        ws.cell(r_idx, 19, s5)
        ws.cell(r_idx, 20, s6)

        # Attended Sessions Formula: =IF(COUNTA(O{r}:T{r})=0,"0 من 0",COUNTIF(O{r}:T{r},"حاضر")&" من "&COUNTA(O{r}:T{r}))
        ws.cell(r_idx, 21, f'=IF(COUNTA(O{r_idx}:T{r_idx})=0,"0 من 0",COUNTIF(O{r_idx}:T{r_idx},"حاضر")&" من "&COUNTA(O{r_idx}:T{r_idx}))')

        # Attendance Rate (%) Formula: =IF(COUNTA(O{r}:T{r})=0,0,COUNTIF(O{r}:T{r},"حاضر")/COUNTA(O{r}:T{r}))
        ws.cell(r_idx, 22, f'=IF(COUNTA(O{r_idx}:T{r_idx})=0,0,COUNTIF(O{r_idx}:T{r_idx},"حاضر")/COUNTA(O{r_idx}:T{r_idx}))')
        ws.cell(r_idx, 22).number_format = "0.0%"

        # Total Score (Tasks + Bonus) Formula: =M{r}+N{r}
        ws.cell(r_idx, 23, f"=M{r_idx}+N{r_idx}")

        # Final Rank Score Formula: =W{r}+(Y{r}*8)
        ws.cell(r_idx, 24, f"=W{r_idx}+(Y{r_idx}*8)")

        # Attended Count: =COUNTIF(O{r}:T{r},"حاضر")
        ws.cell(r_idx, 25, f'=COUNTIF(O{r_idx}:T{r_idx},"حاضر")')

        # Total Sessions: =COUNTA(O{r}:T{r})
        ws.cell(r_idx, 26, f"=COUNTA(O{r_idx}:T{r_idx})")

        # Final Score & Best attendance: =COUNTIF(O{r}:T{r},"حاضر")+(COUNTIF(O{r}:T{r},"مستأذن")*0.5)+X{r}
        ws.cell(r_idx, 27, f'=COUNTIF(O{r_idx}:T{r_idx},"حاضر")+(COUNTIF(O{r_idx}:T{r_idx},"مستأذن")*0.5)+X{r_idx}')

        # Styling row
        for c in range(1, 28):
            cell = ws.cell(r_idx, c)
            cell.font = styles["regular_font"]
            cell.border = styles["grid_border"]
            if c in (1, 3, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27):
                cell.alignment = styles["align_center"]
            else:
                cell.alignment = styles["align_left"]

        # Calculate python-side rank score for Score Board Top 15 display
        present_count = sum(1 for s in (s1, s2, s3, s4, s5, s6) if s == "حاضر")
        excused_count = sum(1 for s in (s1, s2, s3, s4, s5, s6) if s == "مستأذن")
        tot_tasks = t1 + t2 + t3
        final_rank_calc = (tot_tasks + bonus) + (present_count * 8) + present_count + (excused_count * 0.5)
        student_scores.append({
            "id": student_code,
            "name": u.name,
            "score": round(final_rank_calc, 1)
        })

    # Fill Top 15 Leaderboard on Source_Grades (Rows 4 to 18, Columns AC, AD, AE)
    student_scores.sort(key=lambda x: x["score"], reverse=True)
    top_15 = student_scores[:15]
    for idx, item in enumerate(top_15):
        r = 4 + idx
        c_id = ws.cell(r, 29, item["id"])
        c_name = ws.cell(r, 30, item["name"])
        c_score = ws.cell(r, 31, item["score"])
        for c in (c_id, c_name, c_score):
            c.font = styles["bold_font"]
            c.border = styles["grid_border"]
            c.alignment = styles["align_center"]

    # Set Column Widths
    col_widths = {
        "A": 14, "B": 32, "C": 14, "D": 22, "E": 35, "F": 30, "G": 30, "H": 15, "I": 24,
        "J": 10, "K": 10, "L": 10, "M": 16, "N": 12, "O": 10, "P": 10, "Q": 10, "R": 10,
        "S": 10, "T": 10, "U": 16, "V": 18, "W": 22, "X": 16, "Y": 14, "Z": 14, "AA": 24,
        "AC": 14, "AD": 32, "AE": 12
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width


def _get_predefined_teams_data(students, user_by_id, user_by_seat, template_path: str = None):
    """
    Returns teams list mapping team names, supporters/evaluators, and student members.
    If database users have team_id or assigned supporters, it groups them cleanly.
    Also supplements with the 33 predefined teams from the official SFE Model.
    """
    if template_path is None:
        root_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        candidate_1 = os.path.join(root_dir, TEMPLATE_FILENAME)
        candidate_2 = os.path.join(os.getcwd(), TEMPLATE_FILENAME)
        if os.path.exists(candidate_1):
            template_path = candidate_1
        elif os.path.exists(candidate_2):
            template_path = candidate_2

    # Read the default template if available to keep exact official team groupings
    official_teams = []
    if template_path and os.path.exists(template_path):
        try:
            wb = openpyxl.load_workbook(template_path, data_only=True)
            if "Project_Teams" in wb.sheetnames:
                ws = wb["Project_Teams"]
                current_team = None
                current_supporter = None
                current_members = []
                
                for r in range(2, ws.max_row + 1):
                    t_val = ws.cell(r, 1).value
                    num_val = ws.cell(r, 2).value
                    name_val = ws.cell(r, 3).value
                    id_val = ws.cell(r, 4).value
                    supp_val = ws.cell(r, 9).value
                    
                    if t_val:
                        if current_team and current_members:
                            official_teams.append({
                                "team": current_team,
                                "supporter": current_supporter or "",
                                "members": current_members
                            })
                        current_team = str(t_val).strip()
                        current_members = []
                        if supp_val:
                            current_supporter = str(supp_val).strip()
                    elif supp_val:
                        current_supporter = str(supp_val).strip()

                    if id_val or name_val:
                        stu_id_str = str(id_val).strip() if id_val is not None else ""
                        # Try to resolve latest student details from DB
                        db_user = user_by_id.get(stu_id_str) or user_by_seat.get(stu_id_str)
                        current_members.append({
                            "no": num_val or len(current_members) + 1,
                            "id": stu_id_str,
                            "name": (db_user.name if db_user else name_val) or "",
                            "level": (db_user.academic_level if db_user else ws.cell(r, 5).value) or "Level 1",
                            "program": (db_user.program if db_user else ws.cell(r, 6).value) or "General",
                            "mobile": (db_user.phone if db_user else ws.cell(r, 7).value) or "",
                            "email": (db_user.official_email or db_user.email if db_user else ws.cell(r, 8).value) or ""
                        })
                
                if current_team and current_members:
                    official_teams.append({
                        "team": current_team,
                        "supporter": current_supporter or "",
                        "members": current_members
                    })
        except Exception:
            official_teams = []

    # If no template, generate teams from students list (5 per team)
    if not official_teams:
        chunk_size = 5
        for i in range(0, len(students), chunk_size):
            chunk = students[i:i+chunk_size]
            team_num = (i // chunk_size) + 1
            supp_name = chunk[0].assigned_supporter.name if chunk[0].assigned_supporter else "المشرف العام"
            members = []
            for idx, s in enumerate(chunk, 1):
                members.append({
                    "no": idx,
                    "id": str(s.seat_number or s.id),
                    "name": s.name,
                    "level": s.academic_level or "Level 1",
                    "program": s.program or "General",
                    "mobile": s.phone or "",
                    "email": s.official_email or s.email or ""
                })
            official_teams.append({
                "team": f"team {team_num}",
                "supporter": supp_name,
                "members": members
            })

    return official_teams


def get_all_project_teams(db: Session, template_path: str = None):
    if template_path is None:
        root_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        candidate_1 = os.path.join(root_dir, TEMPLATE_FILENAME)
        candidate_2 = os.path.join(os.getcwd(), TEMPLATE_FILENAME)
        if os.path.exists(candidate_1):
            template_path = candidate_1
        elif os.path.exists(candidate_2):
            template_path = candidate_2

    students = db.query(User).filter(User.role.like("%student%")).all()
    user_by_id = {str(u.id): u for u in students}
    user_by_seat = {str(u.seat_number): u for u in students if u.seat_number}

    return _get_predefined_teams_data(students, user_by_id, user_by_seat, template_path)


def _populate_project_scores_sheet(wb, students, user_by_id, user_by_seat, template_path=None, saved_grades_map=None):
    styles = _get_styles()
    if saved_grades_map is None:
        saved_grades_map = {}

    if "Project_Scores" not in wb.sheetnames:
        ws = wb.create_sheet(title="Project_Scores")
    else:
        ws = wb["Project_Scores"]
        # Clear existing non-header cells if needed
        for r in range(ws.max_row, 1, -1):
            ws.delete_rows(r)

    # Headers
    headers = [
        "Team", "No.", "Student Name", "Student ID",
        "Individual Score\n(max 40)", "Full Project Score\n(max 80)",
        "Project Bonus\n(max 25)", "Total Project Score\n(max 145)", "Attendance"
    ]
    for c_idx, h in enumerate(headers, 1):
        cell = ws.cell(1, c_idx, h)
        cell.font = styles["header_font_white"]
        cell.fill = styles["header_fill_blue"]
        cell.alignment = styles["align_center"]
        cell.border = styles["grid_border"]

    teams_data = _get_predefined_teams_data(students, user_by_id, user_by_seat, template_path)
    
    current_row = 2
    for t in teams_data:
        t_name = t["team"]
        members = t["members"]
        m_count = len(members)
        if m_count == 0:
            continue

        start_r = current_row
        end_r = current_row + m_count - 1

        # Check if team has team-level full project score / bonus saved
        team_full_proj = 0
        team_bonus = 0
        for m in members:
            stu_key = str(m["id"]).strip()
            pg = saved_grades_map.get(stu_key)
            if pg:
                if pg.full_project_score:
                    team_full_proj = pg.full_project_score
                if pg.project_bonus:
                    team_bonus = pg.project_bonus

        for m_idx, m in enumerate(members):
            r = current_row + m_idx
            stu_key = str(m["id"]).strip()
            pg = saved_grades_map.get(stu_key)
            
            # Team Name (in top cell of the team group)
            if m_idx == 0:
                ws.cell(r, 1, t_name)
            
            ws.cell(r, 2, m["no"])
            ws.cell(r, 3, m["name"])
            ws.cell(r, 4, m["id"])

            # Individual Score (Col E) - Yellow input cell (locked to 0 if absent)
            if pg and pg.attendance:
                indiv_val = pg.individual_score if (pg.individual_score is not None) else 0
            else:
                indiv_val = 0
            cell_e = ws.cell(r, 5, indiv_val)
            cell_e.fill = styles["input_fill_yellow"]
            cell_e.font = styles["regular_font"]

            # Full Project Score (Col F) - Yellow input cell (merged per team)
            if m_idx == 0:
                cell_f = ws.cell(r, 6, team_full_proj)
                cell_f.fill = styles["input_fill_yellow"]
                cell_f.font = styles["bold_font"]

            # Project Bonus (Col G) - Yellow input cell (merged per team)
            if m_idx == 0:
                cell_g = ws.cell(r, 7, team_bonus)
                cell_g.fill = styles["input_fill_yellow"]
                cell_g.font = styles["bold_font"]

            # Total Project Score Formula: =IF(OR(E{r}="",F{start_r}="",G{start_r}=""),"",E{r}+F{start_r}+G{start_r})
            ws.cell(r, 8, f'=IF(OR(E{r}="",F{start_r}="",G{start_r}=""),"",E{r}+F{start_r}+G{start_r})')

            # Attendance (Col I)
            att_str = "True" if (pg and pg.attendance) else "False"
            ws.cell(r, 9, att_str)

            for c in range(1, 10):
                cell = ws.cell(r, c)
                cell.border = styles["grid_border"]
                if c in (1, 2, 4, 5, 6, 7, 8, 9):
                    cell.alignment = styles["align_center"]
                else:
                    cell.alignment = styles["align_left"]

        # Merge Team Name, Full Project Score, and Bonus columns across the team rows
        if m_count > 1:
            ws.merge_cells(start_row=start_r, start_column=1, end_row=end_r, end_column=1)
            ws.merge_cells(start_row=start_r, start_column=6, end_row=end_r, end_column=6)
            ws.merge_cells(start_row=start_r, start_column=7, end_row=end_r, end_column=7)

        current_row += m_count

    # Column dimensions
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 20
    ws.column_dimensions["I"].width = 14


def _populate_project_teams_sheet(wb, students, user_by_id, user_by_seat):
    styles = _get_styles()
    if "Project_Teams" not in wb.sheetnames:
        ws = wb.create_sheet(title="Project_Teams")
    else:
        ws = wb["Project_Teams"]
        for r in range(ws.max_row, 1, -1):
            ws.delete_rows(r)

    # Headers
    headers = [
        "Team", "No.", "Student Name", "Student ID", "Academic Level", "Program",
        "Mobile Number", "University Email", "Supporter",
        "Individual Score\n(max 30)", "Full Project Score\n(max 80)",
        "Project Bonus\n(max 25)", "Total Project Score\n(max 135)",
        "Tasks Score\n(max 135)", "Attendance Rate (%)", "Attendance Score\n(out of sessions×20)",
        "Tasks %", "Attendance %", "Project %", "Final Grade %",
        "Attendance ≥60%?", "Success Status"
    ]
    for c_idx, h in enumerate(headers, 1):
        cell = ws.cell(1, c_idx, h)
        cell.font = styles["header_font_white"]
        cell.fill = styles["header_fill_dark"]
        cell.alignment = styles["align_center"]
        cell.border = styles["grid_border"]

    # Score Board Header on Project_Teams (AA3 to AD5)
    ws.merge_cells("AA3:AD4")
    sb_header = ws.cell(3, 27, "Score board (top10)")
    sb_header.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    sb_header.fill = styles["header_fill_blue"]
    sb_header.alignment = styles["align_center"]

    sb_cols = ["Leader ID", "Supporter name", "Project Score", "Team member Individual socre"]
    for c_idx, col_name in enumerate(sb_cols, 27):
        cell = ws.cell(5, c_idx, col_name)
        cell.font = styles["bold_font"]
        cell.fill = styles["header_fill_yellow"]
        cell.alignment = styles["align_center"]
        cell.border = styles["grid_border"]

    teams_data = _get_predefined_teams_data(students, user_by_id, user_by_seat)
    
    current_row = 2
    for t in teams_data:
        t_name = t["team"]
        supporter_name = t["supporter"]
        members = t["members"]
        m_count = len(members)
        if m_count == 0:
            continue

        start_r = current_row
        end_r = current_row + m_count - 1

        for m_idx, m in enumerate(members):
            r = current_row + m_idx
            
            # Team Name (Col A)
            if m_idx == 0:
                ws.cell(r, 1, t_name)
            
            # No., Name, ID, Level, Program, Mobile, Email
            ws.cell(r, 2, m["no"])
            ws.cell(r, 3, m["name"])
            ws.cell(r, 4, m["id"])
            ws.cell(r, 5, m["level"])
            ws.cell(r, 6, m["program"])
            ws.cell(r, 7, m["mobile"])
            ws.cell(r, 8, m["email"])

            # Supporter / Discussion Supervisor (Col I)
            if m_idx == 0:
                ws.cell(r, 9, supporter_name)

            # Lookups for Project Scores from Project_Scores sheet
            # J: Individual Score -> =IFERROR(VLOOKUP(D{r},Project_Scores!$D:$E,2,FALSE()),0)
            ws.cell(r, 10, f"=IFERROR(VLOOKUP(D{r},Project_Scores!$D:$E,2,FALSE()),0)")

            # K: Full Project Score -> =IFERROR(VLOOKUP(D{r},Project_Scores!$D:$F,3,FALSE()),0)
            if m_idx == 0:
                ws.cell(r, 11, f"=IFERROR(VLOOKUP(D{r},Project_Scores!$D:$F,3,FALSE()),0)")

            # L: Project Bonus -> =IFERROR(VLOOKUP(D{r},Project_Scores!$D:$G,4,FALSE()),0)
            if m_idx == 0:
                ws.cell(r, 12, f"=IFERROR(VLOOKUP(D{r},Project_Scores!$D:$G,4,FALSE()),0)")

            # M: Total Project Score Formula: =SUM(J{r},K{start_r},L{start_r})
            ws.cell(r, 13, f"=SUM(J{r},K{start_r},L{start_r})")

            # Lookups from Source_Grades sheet:
            # N: Tasks Score (max 135) -> =IFERROR(VLOOKUP(D{r},Source_Grades!$A:$N,13,FALSE()),0)
            ws.cell(r, 14, f"=IFERROR(VLOOKUP(D{r},Source_Grades!$A:$N,13,FALSE()),0)")

            # O: Attendance Rate (%) -> =IFERROR(VLOOKUP(D{r},Source_Grades!$A:$V,22,FALSE())*100,0)
            ws.cell(r, 15, f"=IFERROR(VLOOKUP(D{r},Source_Grades!$A:$V,22,FALSE())*100,0)")

            # P: Attendance Score (out of sessions*20) -> =IFERROR(VLOOKUP(D{r},Source_Grades!$A:$Z,25,FALSE())*20,0)
            ws.cell(r, 16, f"=IFERROR(VLOOKUP(D{r},Source_Grades!$A:$Z,25,FALSE())*20,0)")

            # Q: Tasks % -> =IF(N{r}=0,0,N{r}/135*100)
            ws.cell(r, 17, f"=IF(N{r}=0,0,N{r}/135*100)")

            # R: Attendance % -> =O{r}
            ws.cell(r, 18, f"=O{r}")

            # S: Project % -> =IF(M{r}="",0,M{r}/135*100)
            ws.cell(r, 19, f'=IF(M{r}="",0,M{r}/135*100)')

            # T: Final Grade % Formula: =Q{r}*0.2 + R{r}*0.3 + S{r}*0.5
            ws.cell(r, 20, f"=Q{r}*0.2 + R{r}*0.3 + S{r}*0.5")

            # U: Attendance >= 60%? -> =IF(R{r}>=60,"نعم","لا")
            ws.cell(r, 21, f'=IF(R{r}>=60,"نعم","لا")')

            # V: Success Status -> =IF(AND(T{r}>=60,R{r}>=60),"ناجح","لم ينجح")
            ws.cell(r, 22, f'=IF(AND(T{r}>=60,R{r}>=60),"ناجح","لم ينجح")')

            # Apply cell styles
            for c in range(1, 23):
                cell = ws.cell(r, c)
                cell.font = styles["regular_font"]
                cell.border = styles["grid_border"]
                if c in (1, 2, 4, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22):
                    cell.alignment = styles["align_center"]
                else:
                    cell.alignment = styles["align_left"]

        # Merge Team, Supporter, Full Project Score, and Bonus cells across team members
        if m_count > 1:
            ws.merge_cells(start_row=start_r, start_column=1, end_row=end_r, end_column=1)
            ws.merge_cells(start_row=start_r, start_column=9, end_row=end_r, end_column=9)
            ws.merge_cells(start_row=start_r, start_column=11, end_row=end_r, end_column=11)
            ws.merge_cells(start_row=start_r, start_column=12, end_row=end_r, end_column=12)

        current_row += m_count

    # Fill Team Score Board (Top 10 Leaders/Teams) on Project_Teams (Rows 6 to 15)
    for idx, t in enumerate(teams_data[:10]):
        r = 6 + idx
        leader_id = t["members"][0]["id"] if t["members"] else ""
        supp_name = t["supporter"]
        c1 = ws.cell(r, 27, leader_id)
        c2 = ws.cell(r, 28, supp_name)
        c3 = ws.cell(r, 29, 0)
        c4 = ws.cell(r, 30, f"=SUMIF(D:D,AA{r},J:J)")
        for c in (c1, c2, c3, c4):
            c.font = styles["bold_font"]
            c.border = styles["grid_border"]
            c.alignment = styles["align_center"]

    # Column Widths
    col_widths = {
        "A": 14, "B": 6, "C": 32, "D": 16, "E": 22, "F": 35, "G": 16, "H": 30,
        "I": 24, "J": 15, "K": 18, "L": 16, "M": 18, "N": 16, "O": 18, "P": 20,
        "Q": 14, "R": 14, "S": 14, "T": 16, "U": 16, "V": 16,
        "AA": 16, "AB": 24, "AC": 16, "AD": 28
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width


def _populate_team_summary_sheet(wb):
    styles = _get_styles()
    if "Team_Summary" not in wb.sheetnames:
        ws = wb.create_sheet(title="Team_Summary")
    else:
        ws = wb["Team_Summary"]
        for r in range(ws.max_row, 2, -1):
            ws.delete_rows(r)

    # Title
    ws.cell(1, 1, "ملخص الفرق (يُحدث بعد إدخال درجات البروجيكت)").font = Font(name="Calibri", size=13, bold=True, color="1F4E79")
    
    # Headers
    headers = ["Team", "عدد الطلاب", "متوسط المعدل النهائي", "عدد الناجحين", "نسبة النجاح"]
    for c_idx, h in enumerate(headers, 1):
        cell = ws.cell(2, c_idx, h)
        cell.font = styles["header_font_white"]
        cell.fill = styles["header_fill_dark"]
        cell.alignment = styles["align_center"]
        cell.border = styles["grid_border"]

    # Add 33 teams summary formulas
    for i in range(1, 34):
        r = 2 + i
        team_name = f"team {i}"
        
        ws.cell(r, 1, team_name).alignment = styles["align_center"]
        
        # عدد الطلاب -> =COUNTIF(Project_Teams!A:A,A{r})
        ws.cell(r, 2, f"=COUNTIF(Project_Teams!A:A,A{r})")

        # متوسط المعدل النهائي -> =IFERROR(AVERAGEIF(Project_Teams!A:A,A{r},Project_Teams!T:T),0)
        ws.cell(r, 3, f"=IFERROR(AVERAGEIF(Project_Teams!A:A,A{r},Project_Teams!T:T),0)")
        ws.cell(r, 3).number_format = "0.0%"

        # عدد الناجحين -> =COUNTIFS(Project_Teams!A:A,A{r},Project_Teams!V:V,"ناجح")
        ws.cell(r, 4, f'=COUNTIFS(Project_Teams!A:A,A{r},Project_Teams!V:V,"ناجح")')

        # نسبة النجاح -> =IF(B{r}=0,0,D{r}/B{r})
        ws.cell(r, 5, f"=IF(B{r}=0,0,D{r}/B{r})")
        ws.cell(r, 5).number_format = "0.0%"

        for c in range(1, 6):
            cell = ws.cell(r, c)
            cell.font = styles["regular_font"]
            cell.border = styles["grid_border"]
            cell.alignment = styles["align_center"]

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16


def build_final_submission_workbook(db: Session) -> io.BytesIO:
    """
    Builds the Official Final Delivery Submission Sheet.
    Includes student info, 6 sessions attendance, 3 tasks scores, project scores (Individual Max 30, Full Project Max 80, Bonus Max 25),
    Attendance %, Tasks %, Project %, Final Grade %, and Success Status (ناجح / لم ينجح).
    Sorted numerically by exact database order (1 to 192).
    """
    styles = _get_styles()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Final_Results"

    # Query all students, sessions, tasks, submissions, attendances, project grades
    all_users = db.query(User).filter(User.role.like("%student%")).all()
    all_users.sort(key=lambda u: int(u.id) if str(u.id).isdigit() else 999999)

    tasks = db.query(Task).order_by(Task.id.asc()).all()
    sessions = db.query(SessionSchedule).order_by(SessionSchedule.id.asc()).all()
    all_submissions = db.query(Submission).all()
    all_attendances = db.query(Attendance).all()
    all_pgs = db.query(ProjectGrade).all()

    # Build maps
    sub_map = {}
    for sub in all_submissions:
        if sub.student_id not in sub_map:
            sub_map[sub.student_id] = {}
        sub_map[sub.student_id][sub.task_id] = max(sub_map[sub.student_id].get(sub.task_id, 0.0), sub.score or 0.0)

    att_map = {}
    for att in all_attendances:
        if att.student_id not in att_map:
            att_map[att.student_id] = {}
        st_val = att.status.value if hasattr(att.status, "value") else str(att.status)
        att_map[att.student_id][att.session_id] = st_val

    pg_map = {}
    for pg in all_pgs:
        pg_map[str(pg.student_id).strip()] = pg

    # Predefined teams mapping
    user_by_id = {str(u.id): u for u in all_users}
    user_by_seat = {str(u.seat_number): u for u in all_users if u.seat_number}
    teams_data = _get_predefined_teams_data(all_users, user_by_id, user_by_seat)

    stu_team_map = {}
    team_scores_map = {}
    for t in teams_data:
        t_name = t.get("team") or t.get("team_name") or ""
        supp = t.get("supporter", "")
        full_proj = 0.0
        proj_bon = 0.0
        for m in t.get("members", []):
            m_id = str(m.get("id", "")).strip()
            stu_team_map[m_id] = {"team_name": t_name, "supporter": supp}
            pg = pg_map.get(m_id)
            if pg:
                if pg.full_project_score: full_proj = pg.full_project_score
                if pg.project_bonus: proj_bon = pg.project_bonus
        team_scores_map[t_name.lower().strip()] = (full_proj, proj_bon)

    # Sheet Title Banner
    ws.merge_cells("A1:AA1")
    title_cell = ws.cell(1, 1, "كشف النتائج وحضور المناقشات المعتمدة للتسليم - SFE Final Submission Sheet")
    title_cell.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    title_cell.fill = styles["header_fill_blue"]
    title_cell.alignment = styles["align_center"]

    # Category Headers Row 3 (Grouped Headers)
    ws.merge_cells("A3:G3")
    c_info = ws.cell(3, 1, "بيانات الطالب الأساسية (Student Information)")
    c_info.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    c_info.fill = styles["header_fill_dark"]
    c_info.alignment = styles["align_center"]

    ws.merge_cells("H3:O3")
    c_att = ws.cell(3, 8, "الحضور والغياب (Attendance - 6 Sessions)")
    c_att.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    c_att.fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    c_att.alignment = styles["align_center"]

    ws.merge_cells("P3:T3")
    c_tasks = ws.cell(3, 16, "التاسكات العملية (Tasks Scores)")
    c_tasks.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    c_tasks.fill = PatternFill(start_color="065F46", end_color="065F46", fill_type="solid")
    c_tasks.alignment = styles["align_center"]

    ws.merge_cells("U3:Y3")
    c_proj = ws.cell(3, 21, "مشروع التخرج والمناقشة (Project Scores)")
    c_proj.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    c_proj.fill = PatternFill(start_color="92400E", end_color="92400E", fill_type="solid")
    c_proj.alignment = styles["align_center"]

    ws.merge_cells("Z3:AA3")
    c_final = ws.cell(3, 26, "حالة النجاح والاجتياز (Success Status)")
    c_final.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    c_final.fill = PatternFill(start_color="4C1D95", end_color="4C1D95", fill_type="solid")
    c_final.alignment = styles["align_center"]

    for col in range(1, 28):
        ws.cell(3, col).border = styles["grid_border"]

    # Detailed Headers Row 4
    headers = [
        # Student Info (1..7)
        "ID", "اسم الطالب", "رقم الجلوس", "المستوى الأكاديمي", "القسم / البرنامج", "الفريق (Team)", "مشرف المناقشة (Supporter)",
        # Attendance (8..15)
        "Session 1", "Session 2", "Session 3", "Session 4", "Session 5", "Session 6 (المناقشة)",
        "عدد السيشنات المحضورة", "نسبة الحضور %",
        # Tasks (16..20)
        "Task 1 (max 45)", "Task 2 (max 45)", "Task 3 (max 45)", "مجموع التاسكات (max 135)", "نسبة التاسكات %",
        # Project (21..25)
        "درجة المناقشة الفردية\n(max 30)", "درجة المشروع الجماعية\n(max 80)", "بونص المشروع\n(max 25)", "إجمالي درجة المشروع\n(max 135)", "نسبة المشروع %",
        # Final Evaluation (26..27)
        "شرط الحضور ≥ 60%؟", "حالة النجاح النهائية\n(Success Status)"
    ]

    for c_idx, h in enumerate(headers, 1):
        cell = ws.cell(4, c_idx, h)
        cell.font = styles["header_font_white"]
        if c_idx <= 7:
            cell.fill = styles["header_fill_dark"]
        elif c_idx <= 15:
            cell.fill = styles["header_fill_blue"]
        elif c_idx <= 20:
            cell.fill = PatternFill(start_color="047857", end_color="047857", fill_type="solid")
        elif c_idx <= 25:
            cell.fill = PatternFill(start_color="B45309", end_color="B45309", fill_type="solid")
        else:
            cell.fill = PatternFill(start_color="6D28D9", end_color="6D28D9", fill_type="solid")
        cell.alignment = styles["align_center"]
        cell.border = styles["grid_border"]

    status_ar = {"present": "حاضر", "absent": "غائب", "excused": "مستأذن"}

    # Data Rows (Row 5 onwards)
    for r_idx, u in enumerate(all_users, 5):
        u_subs = sub_map.get(u.id, {})
        u_atts = att_map.get(u.id, {})
        
        # Project Grade
        s_id = str(u.id).strip()
        seat = str(u.seat_number).strip() if u.seat_number else ""
        pg = pg_map.get(s_id) or (pg_map.get(seat) if seat else None)

        t_info = stu_team_map.get(s_id) or (stu_team_map.get(seat) if seat else None) or {}
        team_name = (pg.team_name if pg else None) or t_info.get("team_name", "غير محدد")
        supporter_name = (u.assigned_supporter.name if u.assigned_supporter else None) or t_info.get("supporter", "المشرف العام")

        t_scores = team_scores_map.get(team_name.lower().strip(), (0.0, 0.0))
        full_proj = pg.full_project_score if (pg and pg.full_project_score) else t_scores[0]
        proj_bon = pg.project_bonus if (pg and pg.project_bonus) else t_scores[1]

        is_present_disc = pg.attendance if pg else False
        indiv_proj = min(30.0, pg.individual_score) if (pg and is_present_disc and pg.individual_score is not None) else 0.0

        # Tasks
        t1 = u_subs.get(1, 0.0) if len(tasks) >= 1 else 0.0
        t2 = u_subs.get(2, 0.0) if len(tasks) >= 2 else 0.0
        t3 = u_subs.get(3, 0.0) if len(tasks) >= 3 else 0.0

        # Sessions
        s1 = status_ar.get(u_atts.get(1, "absent"), "غائب") if len(sessions) >= 1 else "غائب"
        s2 = status_ar.get(u_atts.get(2, "absent"), "غائب") if len(sessions) >= 2 else "غائب"
        s3 = status_ar.get(u_atts.get(3, "absent"), "غائب") if len(sessions) >= 3 else "غائب"
        s4 = status_ar.get(u_atts.get(4, "absent"), "غائب") if len(sessions) >= 4 else "غائب"
        s5 = status_ar.get(u_atts.get(5, "absent"), "غائب") if len(sessions) >= 5 else "غائب"
        s6 = "حاضر" if is_present_disc else (status_ar.get(u_atts.get(6, "absent"), "غائب") if len(sessions) >= 6 else "غائب")

        # 1..7: Student Info
        ws.cell(r_idx, 1, int(u.id) if str(u.id).isdigit() else str(u.id))
        ws.cell(r_idx, 2, u.name or "")
        ws.cell(r_idx, 3, str(u.seat_number or ""))
        ws.cell(r_idx, 4, u.academic_level or "")
        ws.cell(r_idx, 5, u.program or "")
        ws.cell(r_idx, 6, team_name)
        ws.cell(r_idx, 7, supporter_name)

        # 8..13: Sessions 1..6
        ws.cell(r_idx, 8, s1)
        ws.cell(r_idx, 9, s2)
        ws.cell(r_idx, 10, s3)
        ws.cell(r_idx, 11, s4)
        ws.cell(r_idx, 12, s5)
        ws.cell(r_idx, 13, s6)

        # 14..15: Attendance Formulas
        # N: Attended Sessions Count -> =COUNTIF(H{r}:M{r},"حاضر")
        ws.cell(r_idx, 14, f'=COUNTIF(H{r_idx}:M{r_idx},"حاضر")')
        # O: Attendance Rate % -> =N{r}/6
        ws.cell(r_idx, 15, f'=N{r_idx}/6')
        ws.cell(r_idx, 15).number_format = "0.0%"

        # 16..20: Tasks
        ws.cell(r_idx, 16, t1)
        ws.cell(r_idx, 17, t2)
        ws.cell(r_idx, 18, t3)
        # S: Total Tasks Score (max 135) -> =SUM(P{r}:R{r})
        ws.cell(r_idx, 19, f'=SUM(P{r_idx}:R{r_idx})')
        # T: Tasks % -> =S{r}/135
        ws.cell(r_idx, 20, f'=S{r_idx}/135')
        ws.cell(r_idx, 20).number_format = "0.0%"

        # 21..25: Project
        ws.cell(r_idx, 21, indiv_proj)
        ws.cell(r_idx, 22, full_proj)
        ws.cell(r_idx, 23, proj_bon)
        # X: Total Project Score (max 135) -> =U{r}+V{r}+W{r}
        ws.cell(r_idx, 24, f'=U{r_idx}+V{r_idx}+W{r_idx}')
        # Y: Project % -> =X{r}/135
        ws.cell(r_idx, 25, f'=X{r_idx}/135')
        ws.cell(r_idx, 25).number_format = "0.0%"

        # 26..27: Final Status (Attendance criteria >= 60%)
        # Z: Attendance Criteria Met >= 60%? -> =IF(O{r}>=0.6,"نعم (مستوفي)","لا (راسب حضور)")
        ws.cell(r_idx, 26, f'=IF(O{r_idx}>=0.6,"نعم (مستوفي)","لا (راسب حضور)")')

        # AA: Success Status -> =IF(O{r}>=0.6,"ناجح","لم ينجح")
        ws.cell(r_idx, 27, f'=IF(O{r_idx}>=0.6,"ناجح","لم ينجح")')
        ws.cell(r_idx, 27).font = styles["bold_font"]

        # Styles
        for c in range(1, 28):
            cell = ws.cell(r_idx, c)
            cell.border = styles["grid_border"]
            if c in (1, 3, 4, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27):
                cell.alignment = styles["align_center"]
            else:
                cell.alignment = styles["align_left"]

    # Column Widths
    col_widths = {
        "A": 8, "B": 32, "C": 15, "D": 18, "E": 26, "F": 16, "G": 24,
        "H": 12, "I": 12, "J": 12, "K": 12, "L": 12, "M": 20,
        "N": 20, "O": 16,
        "P": 16, "Q": 16, "R": 16, "S": 22, "T": 16,
        "U": 22, "V": 22, "W": 18, "X": 22, "Y": 16,
        "Z": 22, "AA": 22
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # ==================== SHEET 2: ORGANIZERS & CREW ====================
    ws_crew = wb.create_sheet(title="Organizers_Crew")
    
    # Query all crew/organizer users
    all_db_users = db.query(User).all()
    crew_users = [
        u for u in all_db_users 
        if any(r in [p.strip().lower() for p in (u.role.value if hasattr(u.role, 'value') else str(u.role or '')).split(',')] 
               for r in ["admin", "hr", "media", "supporter", "instructor"])
    ]
    crew_users.sort(key=lambda u: int(u.id) if str(u.id).isdigit() else 999999)

    # Crew Title Banner
    ws_crew.merge_cells("A1:Q1")
    crew_title = ws_crew.cell(1, 1, "كشف بيانات وحضور فريق العمل والمنظمين - Organizers & Staff Crew Sheet")
    crew_title.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    crew_title.fill = PatternFill(start_color="312E81", end_color="312E81", fill_type="solid") # Deep Indigo
    crew_title.alignment = styles["align_center"]

    # Category Headers Row 3
    ws_crew.merge_cells("A3:G3")
    c_cinfo = ws_crew.cell(3, 1, "بيانات المنظم / عضو الفريق (Organizer Information)")
    c_cinfo.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    c_cinfo.fill = styles["header_fill_dark"]
    c_cinfo.alignment = styles["align_center"]

    ws_crew.merge_cells("H3:O3")
    c_catt = ws_crew.cell(3, 8, "حضور وغياب السيشنات (Attendance - 6 Sessions)")
    c_catt.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    c_catt.fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    c_catt.alignment = styles["align_center"]

    ws_crew.merge_cells("P3:Q3")
    c_cstat = ws_crew.cell(3, 16, "التقييم والالتزام (Commitment Status)")
    c_cstat.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    c_cstat.fill = PatternFill(start_color="4C1D95", end_color="4C1D95", fill_type="solid")
    c_cstat.alignment = styles["align_center"]

    for col in range(1, 18):
        ws_crew.cell(3, col).border = styles["grid_border"]

    # Detailed Headers Row 4
    crew_headers = [
        "ID", "اسم المنظم / عضو الفريق", "الدور التنظيمي / اللجنة (Role)", "رقم الموبايل", "البريد الجامعي / الرسمي",
        "المستوى الأكاديمي", "القسم / البرنامج",
        "Session 1", "Session 2", "Session 3", "Session 4", "Session 5", "Session 6 (المناقشة)",
        "عدد السيشنات المحضورة", "نسبة الحضور %",
        "شرط الالتزام (≥ 60%)", "حالة الاعتماد (Status)"
    ]

    for c_idx, h in enumerate(crew_headers, 1):
        cell = ws_crew.cell(4, c_idx, h)
        cell.font = styles["header_font_white"]
        if c_idx <= 7:
            cell.fill = styles["header_fill_dark"]
        elif c_idx <= 15:
            cell.fill = styles["header_fill_blue"]
        else:
            cell.fill = PatternFill(start_color="6D28D9", end_color="6D28D9", fill_type="solid")
        cell.alignment = styles["align_center"]
        cell.border = styles["grid_border"]

    role_translation = {
        "admin": "إدارة عليا (Admin)",
        "instructor": "محاضر (Instructor)",
        "supporter": "معيد مشرف / مساعد (Supporter)",
        "hr": "لجنة التنظيم وشؤون الطلاب (HR)",
        "media": "لجنة الإعلام والتغطية (Media)",
        "student": "طالب"
    }

    # Data Rows (Row 5 onwards)
    for r_idx, u in enumerate(crew_users, 5):
        u_atts = att_map.get(u.id, {})
        
        # Translate role
        u_roles = [p.strip().lower() for p in (u.role.value if hasattr(u.role, 'value') else str(u.role or '')).split(',') if p.strip()]
        role_labels = [role_translation.get(r, r) for r in u_roles if r != "student"]
        if not role_labels:
            role_labels = [role_translation.get(r, r) for r in u_roles]
        role_display = " + ".join(role_labels) if role_labels else "عضو تنظيم"

        # Sessions
        s1 = status_ar.get(u_atts.get(1, "absent"), "غائب") if len(sessions) >= 1 else "غائب"
        s2 = status_ar.get(u_atts.get(2, "absent"), "غائب") if len(sessions) >= 2 else "غائب"
        s3 = status_ar.get(u_atts.get(3, "absent"), "غائب") if len(sessions) >= 3 else "غائب"
        s4 = status_ar.get(u_atts.get(4, "absent"), "غائب") if len(sessions) >= 4 else "غائب"
        s5 = status_ar.get(u_atts.get(5, "absent"), "غائب") if len(sessions) >= 5 else "غائب"
        s6 = status_ar.get(u_atts.get(6, "absent"), "غائب") if len(sessions) >= 6 else "غائب"

        # 1..7: Info
        ws_crew.cell(r_idx, 1, int(u.id) if str(u.id).isdigit() else str(u.id))
        ws_crew.cell(r_idx, 2, u.name or "")
        ws_crew.cell(r_idx, 3, role_display)
        ws_crew.cell(r_idx, 4, u.phone or "")
        ws_crew.cell(r_idx, 5, u.official_email or u.email or "")
        ws_crew.cell(r_idx, 6, u.academic_level or "")
        ws_crew.cell(r_idx, 7, u.program or "")

        # 8..13: Sessions
        ws_crew.cell(r_idx, 8, s1)
        ws_crew.cell(r_idx, 9, s2)
        ws_crew.cell(r_idx, 10, s3)
        ws_crew.cell(r_idx, 11, s4)
        ws_crew.cell(r_idx, 12, s5)
        ws_crew.cell(r_idx, 13, s6)

        # 14..15: Attendance Formulas
        # N: Attended Sessions Count -> =COUNTIF(H{r}:M{r},"حاضر")
        ws_crew.cell(r_idx, 14, f'=COUNTIF(H{r_idx}:M{r_idx},"حاضر")')
        # O: Attendance Rate % -> =N{r}/6
        ws_crew.cell(r_idx, 15, f'=N{r_idx}/6')
        ws_crew.cell(r_idx, 15).number_format = "0.0%"

        # 16..17: Commitment & Approval
        # P: Attendance Criteria Met >= 60%? -> =IF(O{r}>=0.6,"مستوفي (نعم)","لا")
        ws_crew.cell(r_idx, 16, f'=IF(O{r_idx}>=0.6,"مستوفي (نعم)","لا")')
        # Q: Status -> =IF(O{r}>=0.6,"معتمد (ملتزم)","غير ملتزم")
        ws_crew.cell(r_idx, 17, f'=IF(O{r_idx}>=0.6,"معتمد (ملتزم)","غير ملتزم")')
        ws_crew.cell(r_idx, 17).font = styles["bold_font"]

        # Styles
        for c in range(1, 18):
            cell = ws_crew.cell(r_idx, c)
            cell.border = styles["grid_border"]
            if c in (1, 4, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17):
                cell.alignment = styles["align_center"]
            else:
                cell.alignment = styles["align_left"]

    # Crew Column Widths
    crew_widths = {
        "A": 12, "B": 32, "C": 30, "D": 16, "E": 30, "F": 18, "G": 22,
        "H": 12, "I": 12, "J": 12, "K": 12, "L": 12, "M": 20,
        "N": 20, "O": 16, "P": 22, "Q": 22
    }
    for col_letter, width in crew_widths.items():
        ws_crew.column_dimensions[col_letter].width = width

    # ==================== SHEET 3: SUMMARY STATISTICS ====================
    ws_sum = wb.create_sheet(title="Summary_Statistics")
    ws_sum.cell(1, 1, "الملخص الإحصائي العام للنتائج والتسليم النهائي").font = Font(name="Calibri", size=14, bold=True, color="1F4E79")
    sum_headers = ["المؤشر", "العدد / النسبة", "ملاحظات"]
    for c_idx, h in enumerate(sum_headers, 1):
        c = ws_sum.cell(3, c_idx, h)
        c.font = styles["header_font_white"]
        c.fill = styles["header_fill_dark"]
        c.alignment = styles["align_center"]

    total_r = len(all_users)
    crew_r = len(crew_users)
    sum_rows = [
        ("إجمالي عدد الطلاب المسجلين", f"=COUNTA(Final_Results!B5:B{4+total_r})", "إجمالي طلاب الدفعة"),
        ("إجمالي الطلاب الناجحين", f'=COUNTIF(Final_Results!AA5:AA{4+total_r},"ناجح")', "استوفى شرط الحضور ≥60%"),
        ("إجمالي الطلاب الراسبين (لم ينجح)", f'=COUNTIF(Final_Results!AA5:AA{4+total_r},"لم ينجح")', "لم يستوفِ شرط الحضور"),
        ("نسبة النجاح العامة للطلاب", f'=B4/B3', "نسبة الناجحين من إجمالي الطلاب"),
        ("عدد الطلاب الحاضرين لمناقشة المشروع (Session 6)", f'=COUNTIF(Final_Results!M5:M{4+total_r},"حاضر")', "الطلاب الذين حضروا جلسة المناقشة"),
        ("عدد الطلاب المستوفين لشرط الحضور (≥ 60%)", f'=COUNTIF(Final_Results!Z5:Z{4+total_r},"نعم (مستوفي)")', "حضروا 4 سيشنات على الأقل"),
        ("إجمالي عدد فريق العمل والمنظمين (Organizers & Crew)", f'=COUNTA(Organizers_Crew!B5:B{4+crew_r})', "إجمالي أعضاء اللجان والمشرفين"),
        ("عدد المنظمين الملتزمين بالحضور (≥ 60%)", f'=COUNTIF(Organizers_Crew!Q5:Q{4+crew_r},"معتمد (ملتزم)")', "أعضاء الفريق مستوفي الحضور")
    ]

    for idx, (label, formula, note) in enumerate(sum_rows, 4):
        c1 = ws_sum.cell(idx, 1, label)
        c2 = ws_sum.cell(idx, 2, formula)
        c3 = ws_sum.cell(idx, 3, note)
        c1.font = styles["bold_font"]
        c2.font = styles["bold_font"]
        c3.font = styles["regular_font"]
        for cell in (c1, c2, c3):
            cell.border = styles["grid_border"]
        if idx in (6,):
            c2.number_format = "0.0%"

    ws_sum.column_dimensions["A"].width = 45
    ws_sum.column_dimensions["B"].width = 20
    ws_sum.column_dimensions["C"].width = 45

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
